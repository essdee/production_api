# Copyright (c) 2023, Essdee and contributors
# For license information, please see license.txt


import json
from collections import OrderedDict
from io import BytesIO
from six import string_types
from operator import itemgetter
from typing import Any, Dict, List, Optional, Tuple, TypedDict, Union

import frappe
from frappe import _
from frappe.query_builder.functions import Coalesce, CombineDatetime
from frappe.utils import cint, date_diff, flt, getdate
from frappe.utils.nestedset import get_descendants_of

from production_api.mrp_stock.report.stock_ageing.stock_ageing import FIFOSlots


# class StockBalanceFilter(TypedDict):
# 	from_date: str
# 	to_date: str
# 	item_group: Optional[str]
# 	item: Optional[str]
# 	warehouse: Optional[str]
# 	lot: Optional[str]
# 	# warehouse_type: Optional[str]
# 	# include_uom: Optional[str]  # include extra info in converted UOM
# 	show_stock_ageing_data: bool
# 	show_variant_attributes: bool


SLEntry = Dict[str, Any]


def parse_list_filter(value) -> List[str]:
	"""Normalize a multi-select filter value to a list.

	Accepts a list (MultiSelectList filter), a JSON array string (filters
	serialized over HTTP) or a plain string (internal callers passing a
	single value). Falsy input returns [] (= filter absent).
	"""
	if not value:
		return []
	if isinstance(value, string_types):
		if value.lstrip().startswith("["):
			try:
				value = json.loads(value)
			except ValueError:
				value = [value]
		else:
			value = [value]
	if not isinstance(value, list):
		value = [value]
	return [v for v in value if v]


def execute(filters=None):
	if not filters:
		filters = {}

	company_currency = "INR"
	include_uom = filters.get("include_uom")
	columns = get_columns(filters)
	items = get_items(filters)
	sle = get_stock_ledger_entries(filters, items)

	if filters.get("show_stock_ageing_data"):
		filters["show_warehouse_wise_stock"] = True
		item_wise_fifo_queue = FIFOSlots(filters, sle).generate()

	# if no stock ledger entry found return
	if not sle:
		return columns, []

	iwb_map = get_item_warehouse_map(filters, sle)
	item_map = get_item_details(items, sle, filters)
	# item_reorder_detail_map = get_item_reorder_details(item_map.keys())

	data = []
	conversion_factors = {}

	_func = itemgetter(1)

	to_date = filters.get("to_date")

	for group_by_key in iwb_map:
		item = group_by_key[0]
		warehouse = group_by_key[1]
		lot = group_by_key[2]
		received_type = group_by_key[3]
		if item_map.get(item):
			qty_dict = iwb_map[group_by_key]
			if (filters.get('remove_zero_balance_item') and qty_dict['bal_qty'] == 0):
				continue
			item_reorder_level = 0
			item_reorder_qty = 0
			# if item + warehouse in item_reorder_detail_map:
			# 	item_reorder_level = item_reorder_detail_map[item + warehouse]["warehouse_reorder_level"]
			# 	item_reorder_qty = item_reorder_detail_map[item + warehouse]["warehouse_reorder_qty"]

			report_data = {
				"currency": company_currency,
				"item": item,
				"warehouse": warehouse,
				"received_type":received_type,
				"lot": lot,
				"reorder_level": item_reorder_level,
				"reorder_qty": item_reorder_qty,
			}
			report_data.update(item_map[item])
			report_data.update(qty_dict)
			if include_uom:
				conversion_factors.setdefault(item, item_map[item].conversion_factor)

			if filters.get("show_stock_ageing_data"):
				fifo_queue = item_wise_fifo_queue[(item, warehouse, lot)].get("fifo_queue")

				stock_ageing_data = {"average_age": 0, "earliest_age": 0, "latest_age": 0}
				if fifo_queue:
					fifo_queue = sorted(filter(_func, fifo_queue), key=_func)
					if not fifo_queue:
						continue

					stock_ageing_data["average_age"] = get_average_age(fifo_queue, to_date)
					stock_ageing_data["earliest_age"] = date_diff(to_date, fifo_queue[0][1])
					stock_ageing_data["latest_age"] = date_diff(to_date, fifo_queue[-1][1])

				report_data.update(stock_ageing_data)
			data.append(report_data)
	# add_additional_uom_columns(columns, data, include_uom, conversion_factors)
	return columns, data

def get_average_age(fifo_queue: List, to_date: str) -> float:
	batch_age = age_qty = total_qty = 0.0
	for batch in fifo_queue:
		batch_age = date_diff(to_date, batch[1])

		if isinstance(batch[0], (int, float)):
			age_qty += batch_age * batch[0]
			total_qty += batch[0]
		else:
			age_qty += batch_age * 1
			total_qty += 1

	return flt(age_qty / total_qty, 2) if total_qty else 0.0

def get_columns(filters):
	"""return columns"""
	columns = [
		{
			"label": _("Item"),
			"fieldname": "item",
			"fieldtype": "Link",
			"options": "Item Variant",
			"width": 150,
		},
		{
			"label": _("Item Name"),
			"fieldname": "item_name",
			"fieldtype": "Link",
			"options": "Item",
			"width": 150
		},
		{
			"label": _("Item Group"),
			"fieldname": "item_group",
			"fieldtype": "Link",
			"options": "Item Group",
			"width": 100,
		},
		{
			"label": _("Lot"),
			"fieldname": "lot",
			"fieldtype": "Link",
			"options": "Lot",
			"width": 100,
		},
		{
			"label": _("Warehouse"),
			"fieldname": "warehouse_name",
			"fieldtype": "Link",
			"options": "Supplier",
			"width": 100,
		},
		{
			"fieldname":"received_type",
			"label":"Received Type",
			"fieldtype":"Link",
			"options":"GRN Item Type",
			"width":"80",
		},
	]

	columns.extend(
		[
			{
				"label": _("Stock UOM"),
				"fieldname": "stock_uom",
				"fieldtype": "Link",
				"options": "UOM",
				"width": 90,
			},
			{
				"label": _("Balance Qty"),
				"fieldname": "bal_qty",
				"fieldtype": "Float",
				"width": 100,
				"convertible": "qty",
			},
			{
				"label": _("Balance Value"),
				"fieldname": "bal_val",
				"fieldtype": "Currency",
				"width": 100,
				"options": "currency",
			},
			{
				"label": _("Opening Qty"),
				"fieldname": "opening_qty",
				"fieldtype": "Float",
				"width": 100,
				"convertible": "qty",
			},
			{
				"label": _("Opening Value"),
				"fieldname": "opening_val",
				"fieldtype": "Currency",
				"width": 110,
				"options": "currency",
			},
			{
				"label": _("In Qty"),
				"fieldname": "in_qty",
				"fieldtype": "Float",
				"width": 80,
				"convertible": "qty",
			},
			{"label": _("In Value"), "fieldname": "in_val", "fieldtype": "Float", "width": 80},
			{
				"label": _("Out Qty"),
				"fieldname": "out_qty",
				"fieldtype": "Float",
				"width": 80,
				"convertible": "qty",
			},
			{"label": _("Out Value"), "fieldname": "out_val", "fieldtype": "Float", "width": 80},
			{
				"label": _("Valuation Rate"),
				"fieldname": "val_rate",
				"fieldtype": "Currency",
				"width": 90,
				"convertible": "rate",
				"options": "currency",
			},
		]
	)

	if filters.get("show_inward_date_split"):
		for idx, column in enumerate(columns):
			if column.get("fieldname") == "bal_val":
				columns.insert(
					idx + 1,
					{
						"label": _("Inward Date Split"),
						"fieldname": "inward_split",
						"fieldtype": "Data",
						"width": 220,
					},
				)
				break

	if filters.get("show_stock_ageing_data"):
		columns += [
			{"label": _("Average Age"), "fieldname": "average_age", "width": 100},
			{"label": _("Earliest Age"), "fieldname": "earliest_age", "width": 100},
			{"label": _("Latest Age"), "fieldname": "latest_age", "width": 100},
		]

	if filters.get("show_variant_attributes"):
		columns += [
			{"label": att_name, "fieldname": att_name, "width": 100}
			for att_name in get_variants_attributes()
		]

	return columns


def apply_conditions(query, filters):
	sle = frappe.qb.DocType("Stock Ledger Entry")

	if not filters.get("from_date"):
		frappe.throw(_("'From Date' is required"))

	if to_date := filters.get("to_date"):
		query = query.where(sle.posting_date <= to_date)
	else:
		frappe.throw(_("'To Date' is required"))

	# if company := filters.get("company"):
	# 	query = query.where(sle.company == company)

	# if filters.get("warehouse"):
	# 	query = apply_warehouse_filter(query, sle, filters)
	# elif warehouse_type := filters.get("warehouse_type"):
	# 	query = (
	# 		query.join(warehouse_table)
	# 		.on(warehouse_table.name == sle.warehouse)
	# 		.where(warehouse_table.warehouse_type == warehouse_type)
	# 	)

	return query


def get_stock_ledger_entries(filters, items: List[str]) -> List[SLEntry]:
	sle = frappe.qb.DocType("Stock Ledger Entry")
	supplier = frappe.qb.DocType("Supplier")

	query = (
		frappe.qb.from_(sle).from_(supplier)
		.select(
			sle.item,
			sle.warehouse,
			supplier.supplier_name.as_("warehouse_name"),
			sle.received_type,
			sle.posting_date,
			sle.qty,
			sle.valuation_rate,
			sle.voucher_type,
			sle.qty_after_transaction,
			sle.stock_value_difference,
			sle.item.as_("name"),
			sle.voucher_no,
			sle.stock_value,
			sle.lot,
		)
		.where((sle.docstatus < 2) & (sle.is_cancelled == 0))
		.where(supplier.name == sle.warehouse)
		.orderby(CombineDatetime(sle.posting_date, sle.posting_time))
		.orderby(sle.creation)
		.orderby(sle.qty)
	)
	
	if warehouse := filters.get("warehouse"):
		w = []
		if isinstance(warehouse, list):
			w = warehouse
		elif isinstance(warehouse, string_types):
			w = [warehouse]
		if len(w)!=0:
			query = query.where(sle.warehouse.isin(w))

	if lots := parse_list_filter(filters.get("lot")):
		query = query.where(sle.lot.isin(lots))
	if received_type := filters.get("received_type"):
		query = query.where(sle.received_type == received_type)	
	if items and len(items)!=0 :
		query = query.where(sle.item.isin(items))

	query = apply_conditions(query, filters)
	return query.run(as_dict=True)


def get_opening_vouchers(to_date):
	opening_vouchers = {"Stock Entry": [], "Stock Reconciliation": []}

	# se = frappe.qb.DocType("Stock Entry")
	sr = frappe.qb.DocType("Stock Reconciliation")

	vouchers_data = (
		frappe.qb.from_(
			# (
			# 	frappe.qb.from_(se)
			# 	.select(se.name, Coalesce("Stock Entry").as_("voucher_type"))
			# 	.where((se.docstatus == 1) & (se.posting_date <= to_date) & (se.is_opening == "Yes"))
			# ) + 
			(
				frappe.qb.from_(sr)
				.select(sr.name, Coalesce("Stock Reconciliation").as_("voucher_type"))
				.where((sr.docstatus == 1) & (sr.posting_date <= to_date) & (sr.purpose == "Opening Stock"))
			)
		).select("voucher_type", "name")
	).run(as_dict=True)

	if vouchers_data:
		for d in vouchers_data:
			opening_vouchers[d.voucher_type].append(d.name)

	return opening_vouchers


def get_item_warehouse_map(filters, sle: List[SLEntry]):
	iwb_map = {}
	from_date = getdate(filters.get("from_date"))
	to_date = getdate(filters.get("to_date"))
	opening_vouchers = get_opening_vouchers(to_date)
	float_precision = cint(frappe.db.get_default("float_precision")) or 3
	show_inward_date_split = cint(filters.get("show_inward_date_split"))
	inward_date_queues = {}
	inward_transfer_buckets = {}

	for d in sle:
		group_by_key = get_group_by_key(d)
		if group_by_key not in iwb_map:
			iwb_map[group_by_key] = frappe._dict(
				{
					"warehouse_name": d.warehouse_name,
					"received_type":d.received_type,
					"opening_qty": 0.0,
					"opening_val": 0.0,
					"in_qty": 0.0,
					"in_val": 0.0,
					"out_qty": 0.0,
					"out_val": 0.0,
					"bal_qty": 0.0,
					"bal_val": 0.0,
					"val_rate": 0.0,
				}
			)
		qty_dict = iwb_map[group_by_key]

		if show_inward_date_split:
			update_inward_date_queue(
				d, group_by_key, inward_date_queues, inward_transfer_buckets, float_precision
			)

		if d.voucher_type == "Stock Reconciliation":
			qty_diff = flt(d.qty_after_transaction) - flt(qty_dict.bal_qty)
		else:
			qty_diff = flt(d.qty)

		value_diff = flt(d.stock_value_difference)

		if d.posting_date < from_date or d.voucher_no in opening_vouchers.get(d.voucher_type, []):
			qty_dict.opening_qty += qty_diff
			qty_dict.opening_val += value_diff

		elif d.posting_date >= from_date and d.posting_date <= to_date:
			if flt(qty_diff, float_precision) >= 0:
				qty_dict.in_qty += qty_diff
				qty_dict.in_val += value_diff
			else:
				qty_dict.out_qty += abs(qty_diff)
				qty_dict.out_val += abs(value_diff)

		qty_dict.val_rate = d.valuation_rate
		qty_dict.bal_qty += qty_diff
		qty_dict.bal_val += value_diff

	if show_inward_date_split:
		for group_by_key, qty_dict in iwb_map.items():
			qty_dict["inward_split"] = format_inward_date_split(
				inward_date_queues.get(group_by_key) or [], float_precision
			)

	iwb_map = filter_items_with_no_transactions(iwb_map, float_precision)
	return iwb_map


def update_inward_date_queue(d, group_by_key, inward_date_queues, inward_transfer_buckets, float_precision):
	"""Maintain a FIFO queue of [qty, posting_date, is_reco] slots per row key.

	Mirrors the mechanics of FIFOSlots (stock_ageing) but keyed on the full
	(item, warehouse, lot, received_type) row key, with a hard reset on
	Stock Reconciliation entries.
	"""
	queue = inward_date_queues.setdefault(group_by_key, [])

	if d.voucher_type == "Stock Reconciliation":
		# stock on hand collapses into a single slot dated the reconciliation
		queue.clear()
		reset_qty = flt(d.qty_after_transaction)
		if flt(reset_qty, float_precision):
			queue.append([reset_qty, d.posting_date, True])
		inward_transfer_buckets.pop(group_by_key, None)
		return

	qty = flt(d.qty)
	if qty > 0:
		transfer_data = inward_transfer_buckets.get(group_by_key, {}).get(d.voucher_no)
		if transfer_data:
			# inward/outward from same voucher: restore previously consumed
			# slices with their original dates/flags (oldest first)
			qty_to_restore = qty
			while qty_to_restore:
				if transfer_data and 0 < transfer_data[0][0] <= qty_to_restore:
					# bucket slice is not enough, restore whole slice
					qty_to_restore -= transfer_data[0][0]
					restored = transfer_data.pop(0)
					push_inward_slot(queue, restored[0], restored[1], restored[2], float_precision)
				elif not transfer_data:
					# transfer bucket is empty, extra incoming qty
					push_inward_slot(queue, qty_to_restore, d.posting_date, False, float_precision)
					qty_to_restore = 0
				else:
					# ample slice qty to consume
					transfer_data[0][0] -= qty_to_restore
					push_inward_slot(
						queue, qty_to_restore, transfer_data[0][1], transfer_data[0][2], float_precision
					)
					qty_to_restore = 0
		else:
			push_inward_slot(queue, qty, d.posting_date, False, float_precision)
	elif qty < 0:
		bucket = inward_transfer_buckets.setdefault(group_by_key, {}).setdefault(d.voucher_no, [])
		qty_to_pop = abs(qty)
		while qty_to_pop:
			slot = queue[0] if queue else None
			if slot and 0 < flt(slot[0]) <= qty_to_pop:
				# consume whole slot
				qty_to_pop -= flt(slot[0])
				bucket.append(queue.pop(0))
			elif not queue:
				# negative stock, no balance but qty yet to consume
				queue.append([-(qty_to_pop), d.posting_date, False])
				bucket.append([qty_to_pop, d.posting_date, False])
				qty_to_pop = 0
			else:
				# ample balance (or negative head), consume from first slot
				slot[0] = flt(slot[0]) - qty_to_pop
				bucket.append([qty_to_pop, slot[1], slot[2]])
				qty_to_pop = 0


def push_inward_slot(queue, qty, posting_date, is_reco, float_precision):
	"""Append a positive slot; neutralize a negative/zero head first and
	merge with the tail slot when it has the same date and reco flag."""
	qty = flt(qty)
	if queue and flt(queue[0][0]) <= 0:
		# neutralize 0/negative stock by adding positive stock
		head = queue[0]
		new_qty = flt(head[0]) + qty
		if flt(new_qty, float_precision) > 0:
			queue.pop(0)
			push_inward_slot(queue, new_qty, posting_date, is_reco, float_precision)
		else:
			# still negative: re-dated to the neutralizing inward, so it no
			# longer represents the reconciliation value — drop the tag
			head[0] = new_qty
			head[1] = posting_date
			head[2] = False
		return
	if queue and queue[-1][1] == posting_date and queue[-1][2] == is_reco:
		queue[-1][0] = flt(queue[-1][0]) + qty
	else:
		queue.append([qty, posting_date, is_reco])


def format_inward_date_split(queue, float_precision):
	"""Render the queue as one 'dd-mm-yyyy: qty' line per date, oldest first.

	Same-voucher restores append old-date slices at the queue tail, so slots
	are aggregated by (date, reco flag) and sorted by date before rendering.
	"""
	buckets = {}
	for slot_qty, posting_date, is_reco in queue:
		key = (posting_date, bool(is_reco))
		buckets[key] = buckets.get(key, 0.0) + flt(slot_qty)

	lines = []
	for posting_date, is_reco in sorted(buckets, key=lambda k: (k[0], not k[1])):
		slot_qty = flt(buckets[(posting_date, is_reco)], float_precision)
		if not slot_qty:
			continue
		qty_str = f"{slot_qty:.{float_precision}f}".rstrip("0").rstrip(".")
		date_str = posting_date.strftime("%d-%m-%Y")
		if is_reco:
			lines.append(f"{date_str} (Reco): {qty_str}")
		else:
			lines.append(f"{date_str}: {qty_str}")
	return "\n".join(lines)


def get_group_by_key(row) -> tuple:
	group_by_key = [row.item, row.warehouse, row.lot, row.received_type]
	return tuple(group_by_key)


def filter_items_with_no_transactions(iwb_map, float_precision: float):
	pop_keys = []
	for group_by_key in iwb_map:
		qty_dict = iwb_map[group_by_key]
		no_transactions = True
		for key, val in qty_dict.items():
			if key != "warehouse_name" and key != "received_type" and key != "inward_split":
				val = flt(val, float_precision)
			qty_dict[key] = val
			if key != "val_rate" and key != "inward_split" and val:
				no_transactions = False

		if no_transactions:
			pop_keys.append(group_by_key)

	for key in pop_keys:
		iwb_map.pop(key)
	return iwb_map


def get_items(filters) -> List[str]:
	"Get items based on item code, item group or brand."
	if item := filters.get("item"):
		if isinstance(item, list):
			return item
		elif isinstance(item, string_types):
			return [item]

	item_filters = {}
	# if item_group := filters.get("item_group"):
	# 	children = get_descendants_of("Item Group", item_group, ignore_permissions=True)
	# 	item_filters["item_group"] = ("in", children + [item_group])
	# if brand := filters.get("brand"):
	# 	item_filters["brand"] = brand
	if parent_items := parse_list_filter(filters.get("parent_item")):
		item_filters["item"] = ("in", parent_items)

	return frappe.get_all("Item Variant", filters=item_filters, pluck="name", order_by=None)


def get_item_details(items: List[str], sle: List[SLEntry], filters):
	item_details = {}
	if not items:
		items = list(set(d.item for d in sle))

	if not items:
		return item_details

	item_table = frappe.qb.DocType("Item")
	item_variant_table = frappe.qb.DocType("Item Variant")

	query = (
		frappe.qb.from_(item_table).from_(item_variant_table)
		.select(
			item_variant_table.name,
			item_table.name.as_('item_name'),
			# item_table.description,
			item_table.item_group,
			item_table.brand,
			item_table.default_unit_of_measure.as_('stock_uom')
		)
		.where(
			(item_variant_table.name.isin(items))
			& (item_table.name == item_variant_table.item)
	 	)
	)

	if parent_items := parse_list_filter(filters.get("parent_item")):
		query = query.where(item_table.name.isin(parent_items))

	result = query.run(as_dict=1)

	for item_table in result:
		item_details.setdefault(item_table.name, item_table)

	if filters.get("show_variant_attributes"):
		variant_values = get_variant_values_for(list(item_details))
		item_details = {k: v.update(variant_values.get(k, {})) for k, v in item_details.items()}

	return item_details

def get_variants_attributes() -> List[str]:
	"""Return all item variant attributes."""
	return frappe.get_all("Item Attribute", pluck="name")


def get_variant_values_for(items):
	"""Returns variant values for items."""
	attribute_map = {}

	attribute_info = frappe.get_all(
		"Item Variant Attribute",
		["parent", "attribute", "attribute_value"],
		{
			"parent": ("in", items),
		},
	)

	for attr in attribute_info:
		attribute_map.setdefault(attr["parent"], {})
		attribute_map[attr["parent"]].update({attr["attribute"]: attr["attribute_value"]})

	return attribute_map


HORIZONTAL_EXPORT_EVENT = "stock_balance_horizontal_export_ready"
HORIZONTAL_EXPORT_JOB = (
	"production_api.mrp_stock.report.stock_balance.stock_balance."
	"generate_horizontal_stock_balance_export"
)
HORIZONTAL_EXPORT_TIMEOUT = 1500
HORIZONTAL_EXPORT_STATUS_TTL = 60 * 60


def _format_horizontal_number(value, precision):
	value = flt(value, precision)
	return f"{value:,.{precision}f}".rstrip("0").rstrip(".") or "0"


def _format_horizontal_detail(row, filters):
	float_precision = cint(frappe.db.get_default("float_precision")) or 3
	currency_precision = cint(frappe.db.get_default("currency_precision")) or 2
	currency = row.get("currency") or "INR"
	lines = [
		row.get("item") or "",
		f"Balance Qty: {_format_horizontal_number(row.get('bal_qty'), float_precision)}",
		"Inward Date Split:",
	]
	inward_lines = (row.get("inward_split") or "").splitlines()
	lines.extend([f"  {line}" for line in inward_lines] or ["  --"])
	valuation_rate = _format_horizontal_number(row.get("val_rate"), currency_precision)
	lines.append(f"Valuation Rate: {currency} {valuation_rate}")

	return "\n".join(lines)


def _get_horizontal_attribute_metadata(data):
	"""Return variant attributes, primary attributes and their configured order."""
	variants = list(OrderedDict.fromkeys(row.get("item") for row in data if row.get("item")))
	parent_items = list(
		OrderedDict.fromkeys(row.get("item_name") for row in data if row.get("item_name"))
	)

	variant_attributes = {}
	if variants:
		for row in frappe.get_all(
			"Item Variant Attribute",
			filters={"parent": ("in", variants)},
			fields=["parent", "attribute", "attribute_value", "idx"],
			order_by="parent asc, idx asc",
		):
			variant_attributes.setdefault(row.parent, OrderedDict())[row.attribute] = row.attribute_value

	primary_attributes = {}
	if parent_items:
		primary_attributes = {
			row.name: row.primary_attribute
			for row in frappe.get_all(
				"Item",
				filters={"name": ("in", parent_items)},
				fields=["name", "primary_attribute"],
				order_by=None,
			)
		}

	item_attributes = {}
	primary_mappings = {}
	if parent_items:
		for row in frappe.get_all(
			"Item Item Attribute",
			filters={"parent": ("in", parent_items)},
			fields=["parent", "attribute", "mapping", "idx"],
			order_by="parent asc, idx asc",
		):
			item_attributes.setdefault(row.parent, []).append(row.attribute)
			if row.attribute == primary_attributes.get(row.parent) and row.mapping:
				primary_mappings[row.parent] = row.mapping

	mapping_values = {}
	mapping_names = list(OrderedDict.fromkeys(primary_mappings.values()))
	if mapping_names:
		for row in frappe.get_all(
			"Item Item Attribute Mapping Value",
			filters={"parent": ("in", mapping_names)},
			fields=["parent", "attribute_value", "idx"],
			order_by="parent asc, idx asc",
		):
			mapping_values.setdefault(row.parent, []).append(row.attribute_value)

	primary_value_order = {
		item: mapping_values.get(mapping, []) for item, mapping in primary_mappings.items()
	}
	return variant_attributes, primary_attributes, item_attributes, primary_value_order


def build_horizontal_stock_balance_data(data, filters=None):
	"""Split Stock Balance into simple Work Order-style tables per Item.

	Lot, Received Type, Warehouse, UOM and non-primary attributes stay as
	normal row columns. Only the current Item's primary values are pivoted
	horizontally, which keeps every table compact and familiar.
	"""
	filters = frappe._dict(filters or {})
	data = data or []
	(
		variant_attributes,
		primary_attributes,
		item_attributes,
		primary_value_order,
	) = _get_horizontal_attribute_metadata(data)

	table_sources = OrderedDict()
	for row in data:
		table_sources.setdefault(row.get("item_name"), []).append(row)

	tables = []
	for parent_item, table_data in table_sources.items():
		primary_attribute = primary_attributes.get(parent_item) or None
		common_attributes = [
			attribute
			for attribute in item_attributes.get(parent_item, [])
			if attribute != primary_attribute
		]

		if primary_attribute:
			present_values = []
			for row in table_data:
				value = (
					variant_attributes.get(row.get("item"), {}).get(primary_attribute)
					or "Unspecified"
				)
				if value not in present_values:
					present_values.append(value)
			primary_values = [
				value
				for value in primary_value_order.get(parent_item, [])
				if value in present_values
			]
			primary_values.extend(value for value in present_values if value not in primary_values)
			primary_columns = [(primary_attribute, value) for value in primary_values]
			primary_headers = primary_values
		else:
			primary_columns = [(None, "Details")]
			primary_headers = ["Details"]

		row_groups = OrderedDict()
		for row in table_data:
			attributes = variant_attributes.get(row.get("item"), {})
			common_values = [attributes.get(attribute) or "" for attribute in common_attributes]
			group_key = (
				row.get("lot"),
				row.get("received_type"),
				row.get("warehouse"),
				row.get("warehouse_name"),
				row.get("stock_uom"),
				*common_values,
			)
			row_group = row_groups.setdefault(
				group_key,
				{
					"fixed_values": [
						parent_item or "",
						row.get("item_group") or "",
						row.get("lot") or "",
						row.get("warehouse_name") or "",
						row.get("received_type") or "",
						row.get("stock_uom") or "",
						*common_values,
					],
					"values": {},
				},
			)
			primary_value = (
				attributes.get(primary_attribute) or "Unspecified"
				if primary_attribute
				else "Details"
			)
			row_group["values"][(primary_attribute, primary_value)] = _format_horizontal_detail(
				row, filters
			)

		rows = []
		for row_number, row_group in enumerate(row_groups.values(), 1):
			rows.append(
				[row_number, *row_group["fixed_values"]]
				+ [
					row_group["values"].get(primary_column, "")
					for primary_column in primary_columns
				]
			)
		tables.append(
			{
				"item": parent_item or "",
				"fixed_headers": [
					"S.No.",
					"Item",
					"Item Group",
					"Lot",
					"Warehouse",
					"Received Type",
					"Stock UOM",
					*common_attributes,
				],
				"primary_headers": primary_headers,
				"rows": rows,
			}
		)

	return {
		"tables": tables,
	}


def make_horizontal_stock_balance_workbook(export_data, filters=None):
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter

	tables = export_data["tables"]
	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = "Stock Balance Horizontal"
	worksheet.sheet_view.showGridLines = False

	thin_border = Border(
		left=Side(style="thin", color="D1D8DD"),
		right=Side(style="thin", color="D1D8DD"),
		top=Side(style="thin", color="D1D8DD"),
		bottom=Side(style="thin", color="D1D8DD"),
	)
	current_row = 1
	if not tables:
		worksheet.cell(current_row, 1, "No Stock Balance data found for the selected filters.")

	for table_index, table in enumerate(tables):
		headers = table["fixed_headers"] + table["primary_headers"]
		if table_index:
			current_row += 1

		for column_index, header in enumerate(headers, 1):
			cell = worksheet.cell(current_row, column_index, header)
			cell.font = Font(bold=True, color="1F272E")
			cell.fill = PatternFill("solid", fgColor="F8F9FA")
			cell.alignment = Alignment(vertical="center", wrap_text=True)
			cell.border = thin_border
			width = 28 if header in ("Item", "Warehouse") else 17
			if header == "S.No.":
				width = 8
			if column_index > len(table["fixed_headers"]):
				width = 34
			column_letter = get_column_letter(column_index)
			worksheet.column_dimensions[column_letter].width = max(
				worksheet.column_dimensions[column_letter].width or 0, width
			)
		worksheet.row_dimensions[current_row].height = 32
		current_row += 1

		first_detail_column = len(table["fixed_headers"]) + 1
		for table_row_index, values in enumerate(table["rows"]):
			max_lines = 1
			for column_index, value in enumerate(values, 1):
				cell = worksheet.cell(current_row, column_index, value)
				cell.alignment = Alignment(vertical="top", wrap_text=True)
				cell.border = thin_border
				if table_row_index % 2:
					cell.fill = PatternFill("solid", fgColor="F3F6FA")
				if column_index >= first_detail_column:
					max_lines = max(max_lines, str(value or "").count("\n") + 1)
			worksheet.row_dimensions[current_row].height = min(max(48, max_lines * 14), 300)
			current_row += 1

	worksheet.freeze_panes = "A2"
	worksheet.page_setup.orientation = "landscape"
	worksheet.page_setup.fitToWidth = 1
	worksheet.sheet_properties.pageSetUpPr.fitToPage = True
	return workbook


def _parse_horizontal_export_filters(filters=None):
	if isinstance(filters, string_types):
		filters = json.loads(filters)
	return frappe._dict(filters or {})


def _get_horizontal_export_filename(filters):
	from_date = filters.get("from_date") or "start"
	to_date = filters.get("to_date") or "end"
	return f"Stock_Balance_Horizontal_{from_date}_to_{to_date}.xlsx"


def _build_horizontal_workbook_content(filters):
	_columns, data = execute(filters)
	export_data = build_horizontal_stock_balance_data(data, filters)
	workbook = make_horizontal_stock_balance_workbook(export_data, filters)
	xlsx_file = BytesIO()
	workbook.save(xlsx_file)
	return xlsx_file.getvalue()


def _horizontal_export_cache_key(request_id):
	return f"stock_balance_horizontal_export:{request_id}"


def _set_horizontal_export_status(request_id, user, status, **values):
	payload = {"request_id": request_id, "status": status, **values}
	frappe.cache.set_value(
		_horizontal_export_cache_key(request_id),
		payload,
		user=user,
		expires_in_sec=HORIZONTAL_EXPORT_STATUS_TTL,
	)
	return payload


@frappe.whitelist()
def queue_horizontal_download(filters=None):
	"""Queue the expensive report/XLSX work so the web request returns immediately."""
	frappe.has_permission("Stock Ledger Entry", "read", throw=True)
	filters = _parse_horizontal_export_filters(filters)
	request_id = frappe.generate_hash(length=16)
	user = frappe.session.user
	_set_horizontal_export_status(request_id, user, "queued")

	try:
		frappe.enqueue(
			HORIZONTAL_EXPORT_JOB,
			queue="long",
			timeout=HORIZONTAL_EXPORT_TIMEOUT,
			job_id=f"stock-balance-horizontal-{request_id}",
			filters=dict(filters),
			request_id=request_id,
			export_user=user,
		)
	except Exception as error:
		_set_horizontal_export_status(request_id, user, "failed", error=str(error))
		raise

	return {"request_id": request_id, "status": "queued"}


@frappe.whitelist()
def get_horizontal_download_status(request_id):
	"""Return only the current user's queued export status."""
	if not request_id:
		frappe.throw(_("Export request ID is required"))

	status = frappe.cache.get_value(
		_horizontal_export_cache_key(request_id),
		user=frappe.session.user,
		expires=True,
	)
	return status or {"request_id": request_id, "status": "expired"}


def generate_horizontal_stock_balance_export(filters, request_id, export_user):
	"""Background worker: build a private XLSX and notify the requesting user."""
	_set_horizontal_export_status(request_id, export_user, "running")
	try:
		filters = _parse_horizontal_export_filters(filters)
		file_name = _get_horizontal_export_filename(filters)
		xlsx_content = _build_horizontal_workbook_content(filters)
		file_doc = frappe.get_doc(
			{
				"doctype": "File",
				"file_name": file_name,
				"content": xlsx_content,
				"is_private": 1,
			}
		)
		file_doc.save(ignore_permissions=True)
		# The browser can request the file as soon as the ready event arrives.
		frappe.db.commit()

		result = _set_horizontal_export_status(
			request_id,
			export_user,
			"ready",
			file_url=file_doc.file_url,
			file_name=file_doc.file_name,
		)
		frappe.publish_realtime(HORIZONTAL_EXPORT_EVENT, result, user=export_user)
		return result
	except Exception as error:
		frappe.db.rollback()
		result = _set_horizontal_export_status(
			request_id,
			export_user,
			"failed",
			error=str(error) or _("Horizontal export failed"),
		)
		frappe.publish_realtime(HORIZONTAL_EXPORT_EVENT, result, user=export_user)
		raise


@frappe.whitelist()
def download_horizontal(filters=None):
	frappe.has_permission("Stock Ledger Entry", "read", throw=True)
	filters = _parse_horizontal_export_filters(filters)
	frappe.local.response.filename = _get_horizontal_export_filename(filters)
	frappe.local.response.filecontent = _build_horizontal_workbook_content(filters)
	frappe.local.response.type = "binary"


@frappe.whitelist()
def get_stock_balance(filters=None):
	if isinstance(filters, string_types):
		filters = json.loads(filters)
	return execute(filters)
